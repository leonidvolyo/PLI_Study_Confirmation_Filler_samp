#########################################
# DOC FILLING PROGRAM
#########################################

import openpyxl
import docxtpl
import PySimpleGUI as sg

sg.ChangeLookAndFeel('GreenTan')

layout = [
    [sg.Text('Choose files', size=(35, 1))],
    [sg.Text('Your EXCEL file', size=(15, 1), auto_size_text=False, justification='right'),
     sg.InputText('excel file'), sg.FileBrowse()],
    [sg.Text('Your PATTERN file', size=(15, 1), auto_size_text=False, justification='right'),
     sg.InputText('word file'), sg.FileBrowse()],
    [sg.Submit(), sg.Cancel()]
     ]

event, values = sg.Window('Potvrzeni o studiu FILLER', layout).read(close=True)

workbook = openpyxl.load_workbook(values[0])
students_sheet = workbook.active
names = []
nat = []
dob = []
pas = []

for row in range(2, students_sheet.max_row + 1):
        if students_sheet.cell(row, 1).value != "":
           names.append(students_sheet.cell(row, 1).value)
        else:
           continue

for row in range(2, students_sheet.max_row + 1):
       if students_sheet.cell(row, 2).value != "":
           nat.append(students_sheet.cell(row, 2).value)
       else:
           continue

for row in range(2, students_sheet.max_row + 1):
       if students_sheet.cell(row, 3).value != "":
           dob.append(students_sheet.cell(row, 3).value.strftime("%d.%m.%Y"))
       else:
           continue

for row in range(2, students_sheet.max_row + 1):
       if students_sheet.cell(row, 4).value != "":
           pas.append(students_sheet.cell(row, 4).value)
       else:
           continue

if len(names) % 2 != 0:
    names.append(" ")
if len(nat) % 2 != 0:
    nat.append(" ")
if len(dob) % 2 != 0:
    dob.append(" ")
if len(pas) % 2 != 0:
    pas.append(" ")

for i in range(0, len(names), 2):
      doc = docxtpl.DocxTemplate(values[1])
      context = { 'NAME1' : names[i], 'NAME2' : names[i + 1],'NAT1' : nat[i],'NAT2' : nat[i + 1],'PAS1' : pas[i],'PAS2' : pas[i + 1],'DOB1' : dob[i],'DOB2' : dob[i + 1], 'CJ1' : 200200 + i,'CJ2' : 200201 + i}
      doc.render(context)
      doc.save(f"{names[i]}___{names[i+1]}.docx")



